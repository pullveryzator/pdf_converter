import re
import openpyxl
from openpyxl.styles import Alignment
from typing import Tuple

class MarkdownToExcelConverter:
    def __init__(self):
        """
        Инициализация конвертера Markdown в Exel
        """
        self.wrap_alignment = Alignment(wrap_text=False, vertical='top')

    def process_md_to_excel(self, md_path: str, excel_path: str) -> Tuple[bool, int]:
        try:
            with open(md_path, 'r', encoding='utf-8') as f:
                content = f.read()

            tasks = []
            current_part = "Без части"
            current_image = ""

            lines = content.split('\n')
            i = 0
            while i < len(lines):
                line = lines[i].strip()

                if line.startswith('\\section*{ЧАСТЬ'):
                    current_part = re.search(r'ЧАСТЬ (\d+)', line).group(1)
                    i += 1
                    continue
                
                task_match = re.match(r'^([A-ZА-Я]\d+)\s+(.+)$', line)
                if task_match:
                    task_id = task_match.group(1)
                    question = task_match.group(2)

                    if i+1 < len(lines) and lines[i+1].strip().startswith('!['):
                        current_image = re.search(
                            r'!\[.*?\]\((.*?)\)',
                            lines[i+1]).group(1)
                    else:
                        current_image = ""

                    tasks.append({
                        'part': current_part,
                        'task_id': task_id,
                        'question': question,
                        'image': current_image
                    })
                i += 1

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Задачи"

            ws.append(["Часть", "Номер вопроса", "Вопрос", "Рисунок"])

            for task in tasks:
                ws.append([
                    task['part'],
                    task['task_id'],
                    task['question'],
                    task['image'] if task['image'] else ""
                ])
                ws.cell(row=ws.max_row, column=3).alignment = self.wrap_alignment
            
            wb.save(excel_path)
            return True, len(tasks)
            
        except Exception as e:
            print(f"Ошибка: {str(e)}")
            return False, 0